"""
文件处理器
支持多种文件类型的读取、解析、转换
"""

import os
import logging
from pathlib import Path
from typing import Dict, Any, Optional, List
import json

logger = logging.getLogger(__name__)


class FileHandler:
    """文件处理器"""
    
    # 文件类型映射
    TYPE_HANDLERS = {
        'pdf': 'handle_pdf',
        'doc': 'handle_doc',
        'docx': 'handle_docx',
        'xls': 'handle_xls',
        'xlsx': 'handle_xlsx',
        'txt': 'handle_txt',
        'md': 'handle_markdown',
        'html': 'handle_html',
        'csv': 'handle_csv',
        'json': 'handle_json',
        'xml': 'handle_xml',
        'zip': 'handle_archive',
        'tar': 'handle_archive',
        'gz': 'handle_archive',
    }
    
    def __init__(self):
        pass
    
    def get_file_type(self, filepath: Path) -> str:
        """获取文件类型"""
        return filepath.suffix.lower().lstrip('.')
    
    def read(self, filepath: Path, encoding: str = 'utf-8') -> str:
        """
        读取文件内容
        
        Args:
            filepath: 文件路径
            encoding: 文本编码
            
        Returns:
            文件内容
        """
        suffix = self.get_file_type(filepath)
        
        handler_name = self.TYPE_HANDLERS.get(suffix)
        if handler_name and hasattr(self, handler_name):
            return getattr(self, handler_name)(filepath)
        
        # 默认文本处理
        return self.handle_txt(filepath, encoding)
    
    def handle_txt(self, filepath: Path, encoding: str = 'utf-8') -> str:
        """处理纯文本文件"""
        try:
            with open(filepath, 'r', encoding=encoding) as f:
                return f.read()
        except UnicodeDecodeError:
            # 尝试其他编码
            for enc in ['gbk', 'gb2312', 'latin1']:
                try:
                    with open(filepath, 'r', encoding=enc) as f:
                        return f.read()
                except UnicodeDecodeError:
                    continue
        
        # 二进制模式
        with open(filepath, 'rb') as f:
            return f.read().decode('utf-8', errors='ignore')
    
    def handle_markdown(self, filepath: Path) -> str:
        """处理 Markdown 文件"""
        return self.handle_txt(filepath)
    
    def handle_html(self, filepath: Path) -> str:
        """处理 HTML 文件"""
        content = self.handle_txt(filepath)
        
        # 简单的 HTML 标签移除
        import re
        clean = re.sub(r'<[^>]+>', '', content)
        clean = re.sub(r'\s+', ' ', clean)
        
        return clean.strip()
    
    def handle_json(self, filepath: Path) -> str:
        """处理 JSON 文件"""
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return json.dumps(data, ensure_ascii=False, indent=2)
    
    def handle_csv(self, filepath: Path) -> str:
        """处理 CSV 文件"""
        import csv
        lines = []
        
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            for row in reader:
                lines.append('| ' + ' | '.join(row) + ' |')
        
        return '\n'.join(lines)
    
    def handle_xml(self, filepath: Path) -> str:
        """处理 XML 文件"""
        content = self.handle_txt(filepath)
        
        # 简单的 XML 格式化
        import re
        content = re.sub(r'>\s*<', '>\n<', content)
        
        return content
    
    def handle_pdf(self, filepath: Path) -> str:
        """处理 PDF 文件"""
        try:
            import fitz  # PyMuPDF
            
            doc = fitz.open(filepath)
            text_parts = []
            
            for page in doc:
                text_parts.append(page.get_text())
            
            doc.close()
            return '\n\n'.join(text_parts)
        except ImportError:
            logger.warning("PyMuPDF 未安装，无法解析 PDF")
            return f"[PDF文件: {filepath.name}]"
    
    def handle_docx(self, filepath: Path) -> str:
        """处理 Word DOCX 文件"""
        try:
            from docx import Document
            
            doc = Document(filepath)
            paragraphs = [p.text for p in doc.paragraphs]
            return '\n\n'.join(paragraphs)
        except ImportError:
            logger.warning("python-docx 未安装，无法解析 DOCX")
            return f"[Word文件: {filepath.name}]"
    
    def handle_doc(self, filepath: Path) -> str:
        """处理 Word DOC 文件（较旧格式）"""
        try:
            import subprocess
            
            # 尝试使用 LibreOffice 转换为文本
            result = subprocess.run(
                ['soffice', '--headless', '--convert-to', 'txt', str(filepath)],
                capture_output=True,
                text=True,
                timeout=30
            )
            
            txt_path = filepath.with_suffix('.txt')
            if txt_path.exists():
                return txt_path.read_text(encoding='utf-8')
        except Exception as e:
            logger.warning(f"DOC 解析失败: {e}")
        
        return f"[Word文件: {filepath.name}]"
    
    def handle_xlsx(self, filepath: Path) -> str:
        """处理 Excel XLSX 文件"""
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(filepath, data_only=True)
            sheets_data = []
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = []
                
                for row in sheet.iter_rows(values_only=True):
                    rows.append(' | '.join(str(cell) if cell else '' for cell in row))
                
                sheets_data.append(f"### {sheet_name}\n\n" + '\n'.join(rows))
            
            return '\n\n'.join(sheets_data)
        except ImportError:
            logger.warning("openpyxl 未安装，无法解析 XLSX")
            return f"[Excel文件: {filepath.name}]"
    
    def handle_xls(self, filepath: Path) -> str:
        """处理 Excel XLS 文件"""
        try:
            import subprocess
            
            result = subprocess.run(
                ['soffice', '--headless', '--convert-to', 'csv', str(filepath)],
                capture_output=True,
                text=True,
                timeout=30
            )
            
            csv_path = filepath.with_suffix('.csv')
            if csv_path.exists():
                return self.handle_csv(csv_path)
        except Exception as e:
            logger.warning(f"XLS 解析失败: {e}")
        
        return f"[Excel文件: {filepath.name}]"
    
    def handle_archive(self, filepath: Path) -> str:
        """处理压缩包"""
        return f"[压缩包: {filepath.name}, 大小: {filepath.stat().st_size / 1024:.2f}KB]"
    
    def get_metadata(self, filepath: Path) -> Dict[str, Any]:
        """
        获取文件元数据
        
        Args:
            filepath: 文件路径
            
        Returns:
            元数据字典
        """
        stat = filepath.stat()
        
        return {
            'filename': filepath.name,
            'size': stat.st_size,
            'size_mb': round(stat.st_size / 1024 / 1024, 2),
            'extension': filepath.suffix,
            'created_time': stat.st_ctime,
            'modified_time': stat.st_mtime,
            'accessed_time': stat.st_atime,
        }
    
    def extract_from_archive(self, archive_path: Path, target_dir: Path) -> List[Path]:
        """
        从压缩包提取文件
        
        Args:
            archive_path: 压缩包路径
            target_dir: 目标目录
            
        Returns:
            提取的文件列表
        """
        import tarfile
        import zipfile
        
        target_dir.mkdir(parents=True, exist_ok=True)
        extracted = []
        
        suffix = archive_path.suffix.lower()
        
        try:
            if suffix == '.zip':
                with zipfile.ZipFile(archive_path, 'r') as zf:
                    zf.extractall(target_dir)
                    extracted = [target_dir / name for name in zf.namelist()]
            
            elif suffix in ['.tar', '.gz', '.tgz']:
                with tarfile.open(archive_path, 'r:*') as tf:
                    tf.extractall(target_dir)
                    extracted = [target_dir / member.name for member in tf.getmembers()]
            
            else:
                logger.warning(f"不支持的压缩格式: {suffix}")
        except Exception as e:
            logger.error(f"解压失败: {e}")
        
        return extracted
